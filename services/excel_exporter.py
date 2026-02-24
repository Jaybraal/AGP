from typing import List
"""Excel export service using openpyxl."""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from config import REPORTS_DIR
from database.seed import get_config


_AZUL   = "1A5276"
_AZUL_L = "D6EAF8"
_BOLD   = Font(bold=True)
_HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL   = PatternFill("solid", fgColor=_AZUL)
_ALT_FILL      = PatternFill("solid", fgColor=_AZUL_L)
_CENTER        = Alignment(horizontal="center", vertical="center")
_THIN          = Side(style="thin")
_BORDER        = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _encabezado(ws, titulo: str, columnas: List[str]):
    moneda = get_config("moneda_simbolo") or "RD$"
    agencia = get_config("nombre_agencia")

    ws.append([agencia])
    ws.append([titulo])
    ws.append([f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Moneda: {moneda}"])
    ws.append([])

    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"].font = Font(bold=True, size=11)
    ws["A3"].font = Font(italic=True, size=9)

    ws.append(columnas)
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border    = _BORDER
    return header_row + 1  # next data row


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)


def _guardar(wb: Workbook, nombre: str) -> str:
    os.makedirs(REPORTS_DIR, exist_ok=True)
    path = os.path.join(REPORTS_DIR, nombre)
    wb.save(path)
    return path


# ──────────────────────────────────────────────────────────────────
# Export functions
# ──────────────────────────────────────────────────────────────────

def exportar_caja(reporte: dict) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Caja Diaria"

    fecha = reporte.get("fecha", "")
    columnas = ["Recibo", "Hora", "Cliente", "Préstamo", "Cuota#",
                "Capital", "Intereses", "Mora", "Total", "Método", "Referencia"]
    start = _encabezado(ws, f"Reporte de Caja — {fecha}", columnas)

    for i, pago in enumerate(reporte.get("pagos", [])):
        fill = _ALT_FILL if i % 2 else None
        row = [
            pago.get("numero_recibo", ""),
            pago.get("hora_pago", ""),
            pago.get("cliente_nombre", ""),
            pago.get("numero_prestamo", ""),
            "",
            pago.get("monto_capital",   0),
            pago.get("monto_intereses", 0),
            pago.get("monto_mora",      0),
            pago.get("monto_total",     0),
            pago.get("metodo_pago",     ""),
            pago.get("referencia_pago", ""),
        ]
        ws.append(row)
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill
        for cell in ws[ws.max_row]:
            cell.border = _BORDER

    # Totals row
    t = reporte.get("totales", {})
    ws.append([
        "TOTAL", "", "", "", str(t.get("num_pagos", 0)),
        t.get("total_capital", 0), t.get("total_intereses", 0),
        t.get("total_mora", 0), t.get("total_cobrado", 0), "", "",
    ])
    for cell in ws[ws.max_row]:
        cell.font   = _BOLD
        cell.border = _BORDER

    _auto_width(ws)
    return _guardar(wb, f"caja_{fecha}.xlsx")


def exportar_mora(datos: List[dict]) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Mora"

    columnas = ["Préstamo", "Cliente", "Cédula", "Teléfono",
                "Primera Vencida", "Cuotas Vencidas", "Monto Pendiente", "Saldo Capital"]
    start = _encabezado(ws, "Reporte de Mora", columnas)

    for i, row in enumerate(datos):
        fill = _ALT_FILL if i % 2 else None
        ws.append([
            row.get("numero_prestamo", ""),
            row.get("cliente_nombre",  ""),
            row.get("cedula",          ""),
            row.get("telefono_principal", ""),
            row.get("primera_cuota_vencida", ""),
            row.get("cuotas_vencidas",  0),
            row.get("monto_pendiente",  0),
            row.get("saldo_capital",    0),
        ])
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill
        for cell in ws[ws.max_row]:
            cell.border = _BORDER

    _auto_width(ws)
    from datetime import date
    return _guardar(wb, f"mora_{date.today().isoformat()}.xlsx")


def exportar_proyeccion(datos: List[dict]) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Proyección"

    columnas = ["Fecha", "# Cuotas", "Monto Esperado", "Capital Esperado", "Intereses Esperados"]
    start = _encabezado(ws, "Proyección de Cobros", columnas)

    total_monto = total_cap = total_int = 0.0
    for i, row in enumerate(datos):
        fill = _ALT_FILL if i % 2 else None
        monto = row.get("monto_esperado",      0)
        cap   = row.get("capital_esperado",    0)
        inte  = row.get("intereses_esperados", 0)
        ws.append([
            row.get("fecha_vencimiento", ""),
            row.get("num_cuotas", 0),
            monto, cap, inte,
        ])
        total_monto += monto; total_cap += cap; total_int += inte
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill
        for cell in ws[ws.max_row]:
            cell.border = _BORDER

    ws.append(["TOTAL", "", total_monto, total_cap, total_int])
    for cell in ws[ws.max_row]:
        cell.font = _BOLD; cell.border = _BORDER

    _auto_width(ws)
    from datetime import date
    return _guardar(wb, f"proyeccion_{date.today().isoformat()}.xlsx")


def exportar_amortizacion(prestamo: dict, cuotas: List[dict]) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Amortización"

    columnas = ["#", "Fecha Vencimiento", "Cuota Total",
                "Capital", "Intereses", "Saldo Restante", "Estado"]
    start = _encabezado(ws, f"Tabla de Amortización — {prestamo.get('numero_prestamo', '')}", columnas)

    for i, c in enumerate(cuotas):
        fill = _ALT_FILL if i % 2 else None
        ws.append([
            c.get("numero_cuota",      ""),
            c.get("fecha_vencimiento", ""),
            c.get("cuota_total",       0),
            c.get("capital",           0),
            c.get("intereses",         0),
            c.get("saldo_restante",    0),
            c.get("estado",            ""),
        ])
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill
        for cell in ws[ws.max_row]:
            cell.border = _BORDER

    _auto_width(ws)
    num = prestamo.get("numero_prestamo", "prestamo").replace("/", "-")
    return _guardar(wb, f"amortizacion_{num}.xlsx")
